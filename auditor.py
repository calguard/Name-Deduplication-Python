import os
import sys
import threading
import tkinter
from tkinter import filedialog, messagebox
import customtkinter as ctk
from typing import Dict, List, Optional, Tuple

# Optional native splash support (PyInstaller)
try:
    import pyi_splash as _pyi_splash  # type: ignore
except Exception:
    _pyi_splash = None  # type: ignore

import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import fuzz

# Try to import shared config for province profiles and theme colors
try:
    from config import PROVINCE_PROFILES, THEME_COLORS, ThemeColor, ThemeColors  # type: ignore
except Exception:
    PROVINCE_PROFILES = {}
    THEME_COLORS = {}
    class ThemeColor(str):
        BLUE = "blue"
    class ThemeColors:  # minimal fallback
        def __init__(self, fg_color=None, hover_color=None, hyperlink_color=None):
            self.fg_color = fg_color or [None, None]
            self.hover_color = hover_color or [None, None]
            self.hyperlink_color = hyperlink_color or "#FFFFFF"

# -------------------- Minimal UI helpers (standalone) --------------------
class MessageDialog:
    def __init__(self, parent, title: str, message: str, kind: str = "info"):
        if kind == "error":
            messagebox.showerror(title, message, parent=parent)
        elif kind == "warning":
            messagebox.showwarning(title, message, parent=parent)
        else:
            messagebox.showinfo(title, message, parent=parent)


class Tooltip:
    def __init__(self, widget, text: str, delay_ms: int = 500):
        self.widget = widget
        self.text = text
        self.delay_ms = delay_ms
        self._tipwin = None
        self._after_id = None
        widget.bind("<Enter>", self._on_enter)
        widget.bind("<Leave>", self._on_leave)
        widget.bind("<ButtonPress>", self._on_leave)

    def update_text(self, text: str):
        self.text = text or ""
        # Refresh if showing
        if self._tipwin:
            try:
                label = self._tipwin.children.get("tooltip_label")
                if label:
                    label.configure(text=self.text)
            except Exception:
                pass

    def _on_enter(self, _event=None):
        self._schedule()

    def _on_leave(self, _event=None):
        self._unschedule()
        self._hide()

    def _schedule(self):
        self._unschedule()
        self._after_id = self.widget.after(self.delay_ms, self._show)

    def _unschedule(self):
        if self._after_id:
            try:
                self.widget.after_cancel(self._after_id)
            except Exception:
                pass
            self._after_id = None

    def _show(self):
        if self._tipwin or not self.text:
            return
        try:
            x = self.widget.winfo_rootx() + 20
            y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
            self._tipwin = tkinter.Toplevel(self.widget)
            self._tipwin.wm_overrideredirect(True)
            self._tipwin.wm_geometry(f"+{x}+{y}")
            label = ctk.CTkLabel(self._tipwin, text=self.text, fg_color="#333333", text_color="white", corner_radius=6)
            label._name = "tooltip_label"
            label.pack(ipadx=8, ipady=6)
        except Exception:
            self._tipwin = None

    def _hide(self):
        if self._tipwin:
            try:
                self._tipwin.destroy()
            except Exception:
                pass
            self._tipwin = None

# -------------------- Auditor logic (standalone) --------------------
_FIRST_NAME_MIN_SIM = 60
_LAST_NAME_MIN_SIM = 85
_FULL_NAME_MIN_SIM = 75  # for name-only safety gate
_WRATIO_LOW = 90         # flag if below this for fuzzy matches

_SUFFIX_MAP = {
    'jr': 'jr', 'junior': 'jr',
    'ii': 'ii', '2nd': 'ii', '2': 'ii',
    'sr': 'sr', 'senior': 'sr',
    'i': 'i', '1st': 'i', '1': 'i',
    'iii': 'iii', '3rd': 'iii', '3': 'iii',
    'iv': 'iv', '4th': 'iv', '4': 'iv'
}


def _norm(s: Optional[str]) -> str:
    if s is None:
        return ''
    return str(s).strip()


def _norm_name(s: Optional[str]) -> str:
    # Normalize only for comparison, leave originals intact in outputs
    return _norm(s).replace('.', '').replace('  ', ' ').lower()


def _std_suffix(s: Optional[str]) -> str:
    key = _norm(s).lower().replace('.', '')
    return _SUFFIX_MAP.get(key, key)


def _row_is_empty(values: List[Optional[str]]) -> bool:
    return all(v is None or (isinstance(v, str) and v.strip() == '') for v in values)


def _read_section(ws, start_row: int) -> Tuple[str, int, pd.DataFrame]:
    """
    Parse a section starting at a title row like "--- Officials Found ... ---".
    Returns (section_key, next_row_after_section, dataframe)
    """
    title = str(ws.cell(row=start_row, column=1).value)
    if 'Officials Found' in title:
        section_key = 'officials'
    elif 'Linked Records' in title:
        section_key = 'linking'
    elif 'Duplicates Found' in title:
        section_key = 'dedupe'
    else:
        return ('unknown', start_row + 1, pd.DataFrame())

    header_row = start_row + 1
    # Collect headers until trailing None columns end
    headers: List[str] = []
    col = 1
    while True:
        val = ws.cell(row=header_row, column=col).value
        if val is None and col > 5:  # assume headers done if long None tail
            break
        headers.append(str(val) if val is not None else '')
        col += 1
    # Trim trailing empties
    while headers and headers[-1] == '':
        headers.pop()

    data: List[Dict[str, object]] = []
    r = header_row + 1
    while r <= ws.max_row:
        first_cell = ws.cell(row=r, column=1).value
        # Stop at blank row or next section
        if (isinstance(first_cell, str) and first_cell.startswith('--- ')) or _row_is_empty([ws.cell(row=r, column=c).value for c in range(1, len(headers)+1)]):
            break
        row_dict: Dict[str, object] = {}
        for c, h in enumerate(headers, start=1):
            row_dict[h] = ws.cell(row=r, column=c).value
        data.append(row_dict)
        r += 1

    df = pd.DataFrame(data)
    return (section_key, r + 2, df)  # skip the extra spacer rows that writer adds


def load_report_sections(report_path: str) -> Dict[str, pd.DataFrame]:
    wb = load_workbook(report_path, data_only=True)
    if 'Analysis Report' not in wb.sheetnames:
        raise ValueError('Sheet "Analysis Report" not found in report')
    ws = wb['Analysis Report']

    sections: Dict[str, pd.DataFrame] = {}
    r = 1
    while r <= ws.max_row:
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.startswith('--- '):
            key, next_r, df = _read_section(ws, r)
            if key in ('officials', 'linking', 'dedupe') and not df.empty:
                sections[key] = df.fillna('')
            r = next_r
        else:
            r += 1
    return sections


def _full_name(first: str, middle: str, last: str) -> str:
    return ' '.join([_norm(first), _norm(middle), _norm(last)]).strip()


def _name_similarity(first1: str, middle1: str, last1: str, first2: str, middle2: str, last2: str) -> Dict[str, int]:
    fn1, fn2 = _norm_name(first1), _norm_name(first2)
    ln1, ln2 = _norm_name(last1), _norm_name(last2)
    mn1, mn2 = _norm_name(middle1), _norm_name(middle2)

    full1, full2 = f"{fn1} {mn1} {ln1}".strip(), f"{fn2} {mn2} {ln2}".strip()
    return {
        'first_ratio': fuzz.ratio(fn1, fn2),
        'last_ratio': fuzz.ratio(ln1, ln2),
        'full_ratio': fuzz.ratio(full1, full2),
        'w_ratio': fuzz.WRatio(full1, full2),
        'mn_initial_mismatch': int(bool(mn1 and mn2 and mn1[0:1] != mn2[0:1]))
    }


def _flag_pair(u: Dict[str, str], o: Dict[str, str], remark: str) -> List[str]:
    issues: List[str] = []
    # Field incompatibilities
    b1, b2 = _norm(u.get('Birthdate')), _norm(o.get('Birthdate'))
    if b1 and b2 and b1 != b2:
        issues.append('Birthdate mismatch')

    s1, s2 = _norm(u.get('Sex')).upper(), _norm(o.get('Sex')).upper()
    if s1 and s2 and s1 != s2:
        issues.append('Sex mismatch')

    suf1, suf2 = _std_suffix(u.get('Suffix')), _std_suffix(o.get('Suffix'))
    if suf1 and suf2 and suf1 != suf2:
        issues.append('Suffix mismatch')

    city1, city2 = _norm(u.get('City')).lower(), _norm(o.get('City')).lower()

    # Name similarities
    sim = _name_similarity(
        _norm(u.get('First Name')), _norm(u.get('Middle Name')), _norm(u.get('Last Name')),
        _norm(o.get('First Name')), _norm(o.get('Middle Name')), _norm(o.get('Last Name')),
    )

    if sim['first_ratio'] < _FIRST_NAME_MIN_SIM and _norm_name(u.get('First Name')) != _norm_name(o.get('First Name')):
        issues.append(f"First name similarity is only {int(round(sim['first_ratio']))}%.")
    if sim['last_ratio'] < _LAST_NAME_MIN_SIM and _norm_name(u.get('Last Name')) != _norm_name(o.get('Last Name')):
        issues.append(f"Last name similarity is only {int(round(sim['last_ratio']))}%.")
    if sim['mn_initial_mismatch']:
        issues.append('Middle initial does not match.')

    # Name-only heightened guard: if no discriminating fields present and city differs, flag
    if not (b1 and b2) and not (s1 and s2):
        if city1 and city2 and city1 != city2:
            issues.append('Different city while other details are missing.')
        if sim['full_ratio'] < _FULL_NAME_MIN_SIM:
            issues.append(f"Overall name similarity is {int(round(sim['full_ratio']))}%.")

    # Remarks consistency checks
    # Friendlier, non-technical explanations using rounded percentages
    wr = float(sim['w_ratio'])
    if 'Exact' in (remark or ''):
        # For exact, expect very strong similarity (ideally ~98–100%)
        if wr < 98:
            issues.append(f"Marked 'Exact' but name similarity is only {round(wr)}% (expected ~98–100%).")
    elif 'Fuzzy' in (remark or ''):
        if wr < _WRATIO_LOW:
            issues.append(f"Fuzzy match with low name similarity {round(wr)}% (expected ≥{_WRATIO_LOW}%).")

    return issues


def audit_report(report_path: str, output_csv: Optional[str] = None) -> pd.DataFrame:
    """
    Analyze the Excel report and return a DataFrame of suspicious matches.
    Optionally writes a CSV to `output_csv`.
    """
    if not os.path.exists(report_path):
        raise FileNotFoundError(report_path)

    sections = load_report_sections(report_path)
    suspicious_rows: List[Dict[str, object]] = []

    for sec_key, df in sections.items():
        if df.empty:
            continue
        # Ensure expected columns exist gracefully
        expected_cols = ['group_id', 'Row', 'Remarks', 'First Name', 'Middle Name', 'Last Name', 'Suffix', 'Birthdate', 'City', 'Sex', 'Contact Number']
        for col in expected_cols:
            if col not in df.columns:
                df[col] = ''

        # Group by group_id
        try:
            df['group_id'] = pd.to_numeric(df['group_id'], errors='coerce').fillna(-1).astype(int)
        except Exception:
            pass

        for gid, g in df.groupby('group_id'):
            if gid == -1:
                continue
            # Split user vs reference rows
            user_rows = g[g['Row'].astype(str).str.startswith('userfile', na=False)]
            ref_rows = g[~g['Row'].astype(str).str.startswith('userfile', na=False)]
            if user_rows.empty or ref_rows.empty:
                # For dedupe groups with only user rows, compare all pairs among users
                if sec_key == 'dedupe' and len(g) > 1:
                    pairs = []
                    idxs = list(g.index)
                    for i in range(len(idxs)):
                        for j in range(i+1, len(idxs)):
                            pairs.append((g.loc[idxs[i]], g.loc[idxs[j]]))
                    for u_row, o_row in pairs:
                        remark = str(u_row.get('Remarks', '')) or str(o_row.get('Remarks', ''))
                        issues = _flag_pair(u_row, o_row, remark)
                        if issues:
                            suspicious_rows.append({
                                'section': sec_key,
                                'group_id': gid,
                                'user_row': u_row.get('Row', ''),
                                'other_row': o_row.get('Row', ''),
                                'remark': remark,
                                'issues': '; '.join(sorted(set(issues))),
                                # Snapshot key fields for quick review (dedupe pairs)
                                'user_name': _full_name(u_row.get('First Name', ''), u_row.get('Middle Name', ''), u_row.get('Last Name', '')),
                                'other_name': _full_name(o_row.get('First Name', ''), o_row.get('Middle Name', ''), o_row.get('Last Name', '')),
                                'user_birthdate': _norm(u_row.get('Birthdate', '')),
                                'other_birthdate': _norm(o_row.get('Birthdate', '')),
                                'user_city': _norm(u_row.get('City', '')),
                                'other_city': _norm(o_row.get('City', '')),
                                # Keep contact numbers as-is per policy
                                'user_contact': _norm(u_row.get('Contact Number', '')),
                                'other_contact': _norm(o_row.get('Contact Number', '')),
                                'user_sex': _norm(u_row.get('Sex', '')),
                                'other_sex': _norm(o_row.get('Sex', '')),
                            })
                continue

            for _, u_row in user_rows.iterrows():
                for _, o_row in ref_rows.iterrows():
                    remark = str(o_row.get('Remarks', '')) or str(u_row.get('Remarks', ''))
                    issues = _flag_pair(u_row, o_row, remark)
                    if issues:
                        suspicious_rows.append({
                            'section': sec_key,
                            'group_id': gid,
                            'user_row': u_row.get('Row', ''),
                            'other_row': o_row.get('Row', ''),
                            'remark': remark,
                            'issues': '; '.join(sorted(set(issues))),
                            # Snapshot key fields for quick review
                            'user_name': _full_name(u_row.get('First Name', ''), u_row.get('Middle Name', ''), u_row.get('Last Name', '')),
                            'other_name': _full_name(o_row.get('First Name', ''), o_row.get('Middle Name', ''), o_row.get('Last Name', '')),
                            'user_birthdate': _norm(u_row.get('Birthdate', '')),
                            'other_birthdate': _norm(o_row.get('Birthdate', '')),
                            'user_city': _norm(u_row.get('City', '')),
                            'other_city': _norm(o_row.get('City', '')),
                            # Keep contact numbers as-is per policy
                            'user_contact': _norm(u_row.get('Contact Number', '')),
                            'other_contact': _norm(o_row.get('Contact Number', '')),
                            'user_sex': _norm(u_row.get('Sex', '')),
                            'other_sex': _norm(o_row.get('Sex', '')),
                        })

    result_df = pd.DataFrame(suspicious_rows)
    if output_csv:
        # Do not transform phone numbers; write as-is
        result_df.to_csv(output_csv, index=False, encoding='utf-8-sig')
    return result_df


def resource_path(relative_path: str) -> str:
    """Resolve path for PyInstaller bundle or normal run."""
    try:
        base_path = sys._MEIPASS  # type: ignore[attr-defined]
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


class AuditorApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Deduplication Report Auditor")
        ctk.set_appearance_mode("System")

        # Province/theme resolution (DEFAULT_PROVINCE may be injected at build time)
        self._theme_cfg: Optional[ThemeColors] = None
        if 'DEFAULT_PROVINCE' not in globals():
            globals()['DEFAULT_PROVINCE'] = "Oriental Mindoro"
        try:
            # Optional CLI override: --province "Name" or -p "Name"
            try:
                args = sys.argv[1:]
                if "--province" in args:
                    idx = args.index("--province")
                    if idx + 1 < len(args):
                        globals()["DEFAULT_PROVINCE"] = args[idx + 1]
                elif "-p" in args:
                    idx = args.index("-p")
                    if idx + 1 < len(args):
                        globals()["DEFAULT_PROVINCE"] = args[idx + 1]
            except Exception:
                pass

            province = globals().get("DEFAULT_PROVINCE", None)
            if isinstance(province, str) and province in PROVINCE_PROFILES:
                profile = PROVINCE_PROFILES[province]
                # Update window title to include province-specific title
                try:
                    title = str(profile.title).strip()
                    if title:
                        self.title(f"{title} - Auditor")
                except Exception:
                    pass
                # Pull theme by enum
                try:
                    theme_key = profile.theme  # ThemeColor enum
                    if theme_key in THEME_COLORS:
                        self._theme_cfg = THEME_COLORS[theme_key]
                except Exception:
                    self._theme_cfg = None
        except Exception:
            # Fail safe: continue with default theme
            self._theme_cfg = None

        # Fallback: if no province/theme resolved, default to 'blue' theme if available
        try:
            if self._theme_cfg is None and isinstance(THEME_COLORS, dict) and ThemeColor.BLUE in THEME_COLORS:
                self._theme_cfg = THEME_COLORS[ThemeColor.BLUE]
        except Exception:
            pass

        # Try to set window icon if available
        try:
            logo_path = resource_path("logo.ico")
            self.iconbitmap(logo_path)  # This sets the window icon
        except Exception:
            pass

        # Center and size
        width, height = 640, 520
        self.geometry(f"{width}x{height}+{int((self.winfo_screenwidth()-width)/2)}+{int((self.winfo_screenheight()-height)/2)}")
        self.minsize(560, 440)

        # State
        self.report_path = None
        self.output_csv_path = None
        self._running = False

        # Layout
        self.grid_columnconfigure(0, weight=1)
        # Let the log row (row=4) take the flexible space so controls (row=3)
        # keep their natural height and remain visible in default window size.
        try:
            self.grid_rowconfigure(4, weight=1)
            # Keep a reasonable minimum height for the controls row
            self.grid_rowconfigure(3, minsize=90)
        except Exception:
            pass

        self._build_ui()
        # Center using realized size, then close native splash
        try:
            # Ensure layout sizes are computed
            self.update_idletasks()
            w = self.winfo_width()
            h = self.winfo_height()
            if w <= 1 or h <= 1:
                w, h = 640, 520
            sw = self.winfo_screenwidth()
            sh = self.winfo_screenheight()
            # Use integer division midpoint (original behavior)
            x = (sw - w) // 2
            y = (sh - h) // 2
            self.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            pass

        # Ensure first frame is drawn, then close native splash if present
        try:
            if _pyi_splash:
                self.after(0, lambda: _pyi_splash.close())
        except Exception:
            pass

    def _build_ui(self):
        # Title
        header = ctk.CTkLabel(self, text="Audit Suspicious Matches from Excel Report", font=ctk.CTkFont(size=16, weight="bold"))
        header.grid(row=0, column=0, padx=12, pady=(12, 6), sticky="w")

        # Report selector
        rep_frame = ctk.CTkFrame(self)
        rep_frame.grid(row=1, column=0, padx=12, pady=6, sticky="ew")
        rep_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(rep_frame, text="Analysis Report (.xlsx):").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.report_entry = ctk.CTkEntry(rep_frame, placeholder_text="Select the generated analysis .xlsx file", state="normal")
        self.report_entry.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        self.report_entry_tooltip = Tooltip(self.report_entry, "")
        # Enable Run when user types a path manually
        try:
            self.report_entry.bind("<KeyRelease>", lambda _e: self._on_entry_change())
        except Exception:
            pass
        _btn_kwargs = {}
        try:
            if self._theme_cfg:
                fg = (self._theme_cfg.fg_color or [None, None])[0]
                hv = (self._theme_cfg.hover_color or [None, None])[0]
                if fg and hv:
                    _btn_kwargs = {"fg_color": fg, "hover_color": hv}
        except Exception:
            _btn_kwargs = {}
        ctk.CTkButton(rep_frame, text="Browse...", command=self._choose_report, **_btn_kwargs).grid(row=0, column=2, padx=10, pady=10)

        # Output selector
        out_frame = ctk.CTkFrame(self)
        out_frame.grid(row=2, column=0, padx=12, pady=6, sticky="ew")
        out_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(out_frame, text="Output CSV:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.out_entry = ctk.CTkEntry(out_frame, placeholder_text="Where to save suspicious_matches.csv")
        self.out_entry.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        self.out_entry_tooltip = Tooltip(self.out_entry, "")
        # Enable Run when user types a path manually
        try:
            self.out_entry.bind("<KeyRelease>", lambda _e: self._on_entry_change())
        except Exception:
            pass
        ctk.CTkButton(out_frame, text="Save As...", command=self._choose_output, **_btn_kwargs).grid(row=0, column=2, padx=10, pady=10)

        # Controls + progress
        ctrl_frame = ctk.CTkFrame(self, fg_color="transparent")
        ctrl_frame.grid(row=3, column=0, padx=12, pady=(0, 6), sticky="ew")
        ctrl_frame.grid_columnconfigure(0, weight=1)

        self.run_btn = ctk.CTkButton(ctrl_frame, text="Run Auditor", command=self._run_audit, state="disabled", **_btn_kwargs)
        self.run_btn.grid(row=0, column=0, padx=10, pady=(8, 4), sticky="ew")

        self.progress_text = ctk.CTkLabel(ctrl_frame, text="Idle.")
        self.progress_text.grid(row=1, column=0, padx=10, pady=(0, 6), sticky="w")
        # Progress bar themed with province color if available
        _pb_kwargs = {}
        try:
            if self._theme_cfg:
                pc = (self._theme_cfg.fg_color or [None, None])[0]
                if pc:
                    _pb_kwargs = {"progress_color": pc}
        except Exception:
            _pb_kwargs = {}
        self.progress_bar = ctk.CTkProgressBar(ctrl_frame, **_pb_kwargs)
        self.progress_bar.set(0)
        self.progress_bar.grid(row=2, column=0, padx=10, pady=(0, 8), sticky="ew")

        # Log box (match main app font)
        self.log_text = ctk.CTkTextbox(self, state="disabled", wrap="word", font=ctk.CTkFont(family="Courier New", size=13))
        self.log_text.grid(row=4, column=0, padx=12, pady=(0, 12), sticky="nsew")

        # Footer buttons
        foot = ctk.CTkFrame(self, fg_color="transparent")
        foot.grid(row=5, column=0, padx=12, pady=(0, 12), sticky="ew")
        foot.grid_columnconfigure(0, weight=1)

        self.open_output_btn = ctk.CTkButton(foot, text="Open Output CSV", command=self._open_output, state="disabled", **_btn_kwargs)
        self.open_output_btn.grid(row=0, column=0, padx=10, sticky="w")

        # Key bindings
        self.bind("<Return>", self._on_enter)

    # -------------------- UI helpers --------------------
    def _on_entry_change(self):
        """Called when user types directly in entry boxes."""
        try:
            path = self.report_entry.get().strip()
            if path:
                self.report_path = os.path.normpath(path)
                self.report_entry_tooltip.update_text(self.report_path)
        except Exception:
            pass
        try:
            outp = self.out_entry.get().strip()
            if outp:
                self.output_csv_path = os.path.normpath(outp)
                self.out_entry_tooltip.update_text(self.output_csv_path)
        except Exception:
            pass
        self._enable_run_if_ready()
    def _enable_run_if_ready(self):
        enable = bool(self.report_entry.get().strip()) and bool(self.out_entry.get().strip()) and not self._running
        self.run_btn.configure(state=("normal" if enable else "disabled"))

    def _choose_report(self):
        path = filedialog.askopenfilename(title="Select Analysis Report (.xlsx)", filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*")))
        if not path:
            return
        path = os.path.normpath(path)
        self.report_path = path
        self.report_entry.delete(0, "end")
        self.report_entry.insert(0, path)
        self.report_entry_tooltip.update_text(path)

        # Suggest default output next to report
        base = os.path.splitext(os.path.basename(path))[0]
        suggested = os.path.join(os.path.dirname(path), f"suspicious_matches_{base}.csv")
        if not self.out_entry.get().strip():
            self.out_entry.insert(0, suggested)
            self.out_entry_tooltip.update_text(suggested)
        self._enable_run_if_ready()

    def _choose_output(self):
        default_name = self.out_entry.get().strip() or "suspicious_matches.csv"
        initialdir = os.path.dirname(self.report_entry.get().strip()) if self.report_entry.get().strip() else os.getcwd()
        path = filedialog.asksaveasfilename(title="Save Suspicious Matches CSV", defaultextension=".csv", initialdir=initialdir, initialfile=os.path.basename(default_name), filetypes=(("CSV", "*.csv"),))
        if not path:
            return
        path = os.path.normpath(path)
        self.output_csv_path = path
        self.out_entry.delete(0, "end")
        self.out_entry.insert(0, path)
        self.out_entry_tooltip.update_text(path)
        self._enable_run_if_ready()

    def _run_audit(self):
        report = self.report_entry.get().strip()
        output_csv = self.out_entry.get().strip()
        if not report or not output_csv:
            MessageDialog(self, title="Missing Paths", message="Please select the analysis report and the output CSV path.")
            return
        # Validate report path: must be an existing Excel file
        if not os.path.isfile(report) or not report.lower().endswith((".xlsx", ".xls")):
            MessageDialog(self, title="Invalid Report", message="Please select a valid Excel file (.xlsx or .xls) generated by the app.")
            return
        # Normalize/validate output path: ensure it's a CSV file
        if os.path.isdir(output_csv):
            base = os.path.splitext(os.path.basename(report))[0] or "report"
            output_csv = os.path.join(output_csv, f"suspicious_matches_{base}.csv")
        if not output_csv.lower().endswith(".csv"):
            output_csv = output_csv + ".csv"
        # Reflect any normalization to the UI
        try:
            self.out_entry.delete(0, "end")
            self.out_entry.insert(0, output_csv)
            self.out_entry_tooltip.update_text(output_csv)
        except Exception:
            pass
        # Ensure the output directory exists
        try:
            out_dir = os.path.dirname(output_csv) or "."
            os.makedirs(out_dir, exist_ok=True)
        except Exception:
            pass

        # Lock UI
        self._running = True
        self._enable_run_if_ready()
        self.open_output_btn.configure(state="disabled")
        self.progress_text.configure(text="Running auditor...")
        self.progress_bar.configure(mode="indeterminate")
        self.progress_bar.start()
        self._clear_log()
        self._append_log("Starting audit...\n")

        t = threading.Thread(target=self._run_audit_thread, args=(report, output_csv), daemon=True)
        t.start()

    def _run_audit_thread(self, report: str, output_csv: str):
        try:
            df = audit_report(report, output_csv)
            n = 0
            try:
                n = len(df) if df is not None else 0
            except Exception:
                n = 0
            self.after(0, lambda: self._on_audit_complete(success=True, count=n, output_csv=output_csv))
        except PermissionError as e:
            msg = f"Permission error: {e}. If the CSV is open, please close it or choose a new filename."
            self.after(0, lambda m=msg: self._on_audit_error(m))
        except Exception as e:
            msg = f"Unexpected error: {e}"
            self.after(0, lambda m=msg: self._on_audit_error(m))

    def _on_audit_complete(self, success: bool, count: int, output_csv: str):
        self.progress_bar.stop()
        self.progress_bar.configure(mode="determinate")
        self.progress_bar.set(1)
        self.progress_text.configure(text=f"Done. Suspicious rows: {count}")
        self._append_log(f"✅ Done. Wrote suspicious matches to:\n{output_csv}\n")
        self.open_output_btn.configure(state=("normal" if os.path.exists(output_csv) else "disabled"))
        self._running = False
        self._enable_run_if_ready()

    def _on_audit_error(self, message: str):
        self.progress_bar.stop()
        self.progress_bar.configure(mode="determinate")
        self.progress_bar.set(0)
        self.progress_text.configure(text="Error")
        self._append_log("❌ " + message + "\n")
        MessageDialog(self, title="Audit Failed", message=message)
        self._running = False
        self._enable_run_if_ready()

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _append_log(self, text: str):
        def do_append():
            self.log_text.configure(state="normal")
            self.log_text.insert("end", text)
            self.log_text.configure(state="disabled")
            self.log_text.see("end")
        # Always marshal to UI thread
        self.after(0, do_append)

    def _open_output(self):
        path = self.out_entry.get().strip()
        if not path or not os.path.exists(path):
            MessageDialog(self, title="Not Found", message="Output CSV not found.")
            return
        try:
            if sys.platform == "win32":
                os.startfile(os.path.normpath(path))  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                os.system(f'open "{os.path.normpath(path)}"')
            else:
                os.system(f'xdg-open "{os.path.normpath(path)}"')
        except Exception as e:
            MessageDialog(self, title="Open Failed", message=f"Could not open file: {e}")

    def _on_enter(self, event=None):
        if self.run_btn.cget("state") == "normal":
            try:
                self.run_btn.invoke()
            except Exception:
                pass


if __name__ == "__main__":
    app = AuditorApp()
    app.mainloop()
