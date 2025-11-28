import pandas as pd
import os
import re
import logging
import warnings
from collections import defaultdict
from datetime import datetime
from multiprocessing import Pool, cpu_count
from functools import partial
import itertools

from rapidfuzz import fuzz
import jellyfish
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, DoughnutChart, Reference, PieChart
from openpyxl.chart.series import DataPoint # THIS IS THE CORRECT IMPORT
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

from typing import Dict, Any

import config
import excel_converter
from config import THEME_COLORS, PROVINCE_PROFILES

INTENDED_COLS = ["First Name", "Middle Name", "Last Name", "Suffix", "Birthdate", "City", "Sex", "Contact Number"]

def get_automatic_output_path(user_filepath, province_name, is_pdf):
    output_dir = os.path.dirname(user_filepath)
    base_name = os.path.splitext(os.path.basename(user_filepath))[0]
    
    file_ext = ".pdf" if is_pdf else ".xlsx"
    province_tag = province_name.replace(" ", "_")
    
    output_base = os.path.join(output_dir, f"{base_name}_{province_tag}_report")
    
    # Always start numbering at 1 to avoid having an unsuffixed first file
    counter = 1
    final_path = f"{output_base}_{counter}{file_ext}"
    while os.path.exists(final_path):
        counter += 1
        final_path = f"{output_base}_{counter}{file_ext}"
    return final_path

def create_summary_section(summary_stats, user_df, master_df, officials_df):
    # Get the actual counts from summary_stats
    officials_count = summary_stats.get("officials", 0)
    linking_count = summary_stats.get("linking", 0)
    duplicates_count = summary_stats.get("duplicates", 0)
    unique_count = summary_stats.get("unique", 0)
    total_user_records = summary_stats.get("total", 0)
    
    # Calculate total scanned based on your old script logic:
    # Officials: user file records + officials database records
    officials_scanned = len(user_df) + len(officials_df) if officials_df is not None and not officials_df.empty else len(user_df)
    
    # Linked Records: user file records + master database records
    linking_scanned = len(user_df) + len(master_df) if master_df is not None and not master_df.empty else len(user_df)
    
    # Duplicates: user file records (scanned against itself)
    duplicates_scanned = len(user_df)
    
    # Unique: no total scanned (just the count of unique records found)
    unique_scanned = ""
    
    # Total: sum of all records processed across all databases
    total_scanned = len(user_df) + (len(officials_df) if officials_df is not None and not officials_df.empty else 0) + (len(master_df) if master_df is not None and not master_df.empty else 0)
    
    table_data = [
        ["Officials (User File)", officials_scanned, officials_count],
        ["Linked Records (User+Master)", linking_scanned, linking_count],
        ["Duplicates (User File)", duplicates_scanned, duplicates_count],
        ["Unique Records", unique_scanned, unique_count],
        ["Total", total_scanned, total_user_records]
    ]
    return table_data

def format_duration(seconds):
    if seconds < 1: return "less than 1 second"
    if seconds < 60: return f"{int(seconds)} second(s)"
    minutes, seconds = divmod(int(seconds), 60)
    if minutes < 60: return f"{minutes} minute(s), {seconds} second(s)"
    hours, minutes = divmod(minutes, 60)
    return f"{hours} hour(s), {minutes} minute(s), {seconds} second(s)"

def _precompute_dataframe(df, symmetrical_map):
    if df is None or df.empty: return df
    df['_opt_mname_raw'] = df['Middle Name'].fillna('').astype(str)
    df['_opt_lname_raw'] = df['Last Name'].fillna('').astype(str)
    df['_opt_fname_exp'] = df['First Name'].fillna('').astype(str).str.replace(r'^\b(Ma\.|Ma)\b', 'Maria', regex=True, flags=re.IGNORECASE)
    df['_opt_fname_std'] = df['_opt_fname_exp'].str.lower().str.replace('.', '', regex=False).str.replace(' ', '', regex=False)
    df['_opt_nickname_set'] = df['_opt_fname_std'].apply(lambda x: symmetrical_map.get(x, {x}))
    df['_opt_soundex_lname'] = df['_opt_lname_raw'].apply(jellyfish.soundex)
    suffix_map = {'jr': 'jr', 'junior': 'jr', 'ii': 'ii', '2nd': 'ii', '2': 'ii', 'sr': 'sr', 'senior': 'sr', 'i': 'i', '1st': 'i', '1': 'i', 'iii': 'iii', '3rd': 'iii', '3': 'iii', 'iv': 'iv', '4th': 'iv', '4': 'iv'}
    s_series = df['Suffix'].fillna('').astype(str).str.lower().str.replace('.', '', regex=False).str.strip()
    df['_opt_suffix_std'] = s_series.map(suffix_map).fillna(s_series)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        df['_opt_bdate_std'] = pd.to_datetime(df['Birthdate'], errors='coerce').dt.strftime('%Y-%m-%d').fillna('')
    # City is already normalized in main.py via normalize_city(); just lowercase and trim
    df['_opt_city_std'] = df['City'].fillna('').astype(str).str.lower().str.strip()
    return df

def _calculate_match_confidence_optimized(rec1, rec2):
    sex1, sex2 = str(rec1.get("Sex", "")).upper(), str(rec2.get("Sex", "")).upper()
    if sex1 and sex2 and sex1 != sex2: return -1000
    if rec1['_opt_suffix_std'] and rec2['_opt_suffix_std'] and rec1['_opt_suffix_std'] != rec2['_opt_suffix_std']: return -1000
    fn1, fn2 = rec1['_opt_fname_exp'], rec2['_opt_fname_exp']
    mn1, mn2 = rec1['_opt_mname_raw'], rec2['_opt_mname_raw']
    ln1, ln2 = rec1['_opt_lname_raw'], rec2['_opt_lname_raw']
    common_names = rec1['_opt_nickname_set'].intersection(rec2['_opt_nickname_set'])
    if common_names:
        formal_name = max(common_names, key=len).capitalize()
        fn1, fn2 = formal_name, formal_name
    fn_score, mn_score = fuzz.ratio(fn1, fn2), fuzz.ratio(mn1, mn2)
    full_name1, full_name2 = f"{fn1} {mn1} {ln1}".strip().replace("  ", " "), f"{fn2} {mn2} {ln2}".strip().replace("  ", " ")
    token_set_score = fuzz.token_set_ratio(full_name1, full_name2)
    confidence_score = 0.0
    is_phonetic_match = (jellyfish.soundex(fn1) == jellyfish.soundex(fn2) and rec1['_opt_soundex_lname'] == rec2['_opt_soundex_lname'])
    if is_phonetic_match and fn_score > 80: confidence_score += 40
    if fuzz.WRatio(full_name1, full_name2) > 95: confidence_score += 30
    bdate1, bdate2 = rec1['_opt_bdate_std'], rec2['_opt_bdate_std']
    if bdate1 and bdate2:
        if bdate1 == bdate2: confidence_score += 100
        else: confidence_score -= 150
    confidence_score += token_set_score
    if rec1['_opt_city_std'] and rec2['_opt_city_std'] and rec1['_opt_city_std'] != rec2['_opt_city_std']: confidence_score -= 30
    if mn1 and mn2:
        if len(mn1) > 1 and len(mn2) > 1 and mn_score < 65: confidence_score -= 80
        elif mn1[0].lower() != mn2[0].lower(): confidence_score -= 60
    return confidence_score

def _calculate_adaptive_match_confidence(rec1, rec2):
    """
    Enhanced confidence calculation for datasets without birthdate/sex.
    Uses much stricter name similarity requirements for name-only matching.
    """
    # Start with base score from optimized algorithm
    base_score = _calculate_match_confidence_optimized(rec1, rec2)
    
    # If base score is already a hard reject, return it
    if base_score <= -1000:
        return base_score
    
    # Check if we have discriminating fields available
    has_birthdate_raw = bool(str(rec1.get("Birthdate", "")).strip() and str(rec2.get("Birthdate", "")).strip())
    has_birthdate_processed = bool(rec1['_opt_bdate_std'] and rec2['_opt_bdate_std'])
    has_sex = bool(str(rec1.get("Sex", "")).strip() and str(rec2.get("Sex", "")).strip())
    
    # If we have birthdate data, use original algorithm (sex alone isn't discriminating enough)
    if has_birthdate_raw or has_birthdate_processed:
        return base_score
    
    # For name-only matching, apply MUCH stricter criteria
    fn1, fn2 = rec1['_opt_fname_exp'], rec2['_opt_fname_exp']
    ln1, ln2 = rec1['_opt_lname_raw'], rec2['_opt_lname_raw']
    mn1, mn2 = rec1['_opt_mname_raw'], rec2['_opt_mname_raw']
    
    # STRICT RULE 1: First names must be reasonably similar (60%+) OR exact match
    fn_similarity = fuzz.ratio(fn1, fn2)
    
    if fn_similarity < 60 and fn1.lower() != fn2.lower():
        return -1000  # Hard reject for insufficient first name similarity
    
    # STRICT RULE 2: Last names must be similar (85%+) OR exact match  
    ln_similarity = fuzz.ratio(ln1, ln2)
    if ln_similarity < 85 and ln1.lower() != ln2.lower():
        return -1000  # Hard reject for insufficient last name similarity
    
    # STRICT RULE 3: If middle names exist, they should be compatible
    if mn1 and mn2 and len(mn1) > 1 and len(mn2) > 1:
        mn_similarity = fuzz.ratio(mn1, mn2)
        if mn_similarity < 80 and mn1[0].lower() != mn2[0].lower():
            return -1000  # Hard reject for incompatible middle names
    
    # STRICT RULE 4: Full name similarity must be very high (95%+)
    full_name1 = f"{fn1} {mn1} {ln1}".strip().replace("  ", " ")
    full_name2 = f"{fn2} {mn2} {ln2}".strip().replace("  ", " ")
    full_similarity = fuzz.ratio(full_name1, full_name2)
    
    if full_similarity < 75:
        return -1000  # Hard reject for insufficient overall similarity
    
    # Apply conservative penalty for name-only matches
    penalty_factor = 0.9  # Reduce confidence by 10%
    
    return int(base_score * penalty_factor)

def test_name_pair(first1, last1, first2, last2, middle1="", middle2=""):
    """Quick test function to check if two names would match with current rules"""
    import jellyfish
    
    # Create mock records with all required fields
    rec1 = {
        '_opt_fname_exp': first1,
        '_opt_lname_raw': last1,
        '_opt_mname_raw': middle1,
        '_opt_bdate_std': None,
        '_opt_suffix_std': '',
        '_opt_nickname_set': set(),
        '_opt_soundex_lname': jellyfish.soundex(last1) if last1 else '',
        '_opt_city_std': '',
        'Sex': 'Female',
        'Birthdate': ''
    }
    rec2 = {
        '_opt_fname_exp': first2,
        '_opt_lname_raw': last2,
        '_opt_mname_raw': middle2,
        '_opt_bdate_std': None,
        '_opt_suffix_std': '',
        '_opt_nickname_set': set(),
        '_opt_soundex_lname': jellyfish.soundex(last2) if last2 else '',
        '_opt_city_std': '',
        'Sex': 'Female',
        'Birthdate': ''
    }
    
    score = _calculate_adaptive_match_confidence(rec1, rec2)
    threshold = 75  # Current threshold (110 - 35)
    
    print(f"\n=== TESTING: '{first1} {last1}' vs '{first2} {last2}' ===")
    print(f"Score: {score}")
    print(f"Threshold: {threshold}")
    print(f"Result: {'MATCH' if score > threshold else 'NO MATCH'}")
    
    return score > threshold

def compare_records_strict_optimized(rec1: Dict[str, Any], rec2: Dict[str, Any]) -> str:
    """Compare two records using optimized strict matching.
    
    Args:
        rec1: First record to compare as a dictionary with string keys
        rec2: Second record to compare as a dictionary with string keys
        
    Returns:
        str: "Exact Match" if score > 198, otherwise "No Match"
    """
    score = _calculate_match_confidence_optimized(rec1, rec2)
    return "Exact Match" if score > 198 else "No Match"

def compare_records_standard_optimized(rec1: Dict[str, Any], rec2: Dict[str, Any]) -> str:
    """Compare two records using optimized standard matching.
    
    Args:
        rec1: First record to compare as a dictionary with string keys
        rec2: Second record to compare as a dictionary with string keys
        
    Returns:
        str: "Fuzzy Match" if score > 110, otherwise "No Match"
    """
    if _calculate_match_confidence_optimized(rec1, rec2) > 110: 
        return "Fuzzy Match"
    return "No Match"

def compare_records_lenient_optimized(rec1: Dict[str, Any], rec2: Dict[str, Any]) -> str:
    """Compare two records using optimized lenient matching.
    
    Args:
        rec1: First record to compare as a dictionary with string keys
        rec2: Second record to compare as a dictionary with string keys
        
    Returns:
        str: "Fuzzy Match" if score > 95, otherwise "No Match"
    """
    score = _calculate_match_confidence_optimized(rec1, rec2)
    return "Fuzzy Match" if score > 95 else "No Match"

def compare_records_strict_adaptive(rec1: Dict[str, Any], rec2: Dict[str, Any]) -> str:
    """Adaptive strict comparison that maintains accuracy regardless of available fields.
    
    Args:
        rec1: First record to compare as a dictionary with string keys
        rec2: Second record to compare as a dictionary with string keys
        
    Returns:
        str: "Exact Match" if score > 198, otherwise "No Match"
    """
    score = _calculate_adaptive_match_confidence(rec1, rec2)
    return "Exact Match" if score > 198 else "No Match"

def compare_records_standard_adaptive(rec1: Dict[str, Any], rec2: Dict[str, Any]) -> str:
    """Adaptive standard comparison that maintains accuracy regardless of available fields.
    
    Args:
        rec1: First record to compare as a dictionary with string keys
        rec2: Second record to compare as a dictionary with string keys
        
    Returns:
        str: "Fuzzy Match" if score > 110, otherwise "No Match"
    """
    score = _calculate_adaptive_match_confidence(rec1, rec2)
    return "Fuzzy Match" if score > 110 else "No Match"

def compare_records_lenient_adaptive(rec1: Dict[str, Any], rec2: Dict[str, Any]) -> str:
    """Adaptive lenient comparison that maintains accuracy regardless of available fields.
    
    Args:
        rec1: First record to compare as a dictionary with string keys
        rec2: Second record to compare as a dictionary with string keys
        
    Returns:
        str: "Fuzzy Match" if score > 95, otherwise "No Match"
    """
    score = _calculate_adaptive_match_confidence(rec1, rec2)
    return "Fuzzy Match" if score > 95 else "No Match"

def compare_records_strict_configurable(rec1: Dict[str, Any], rec2: Dict[str, Any]) -> bool:
    """
    Compare two records using strict configurable thresholds.
    
    Args:
        rec1: First record to compare as a dictionary with string keys
        rec2: Second record to compare as a dictionary with string keys
    
    Returns:
        bool: True if records match according to strict criteria, False otherwise
    """
    # Use enhanced adaptive algorithm when adaptive mode is enabled
    if config.ADAPTIVE_MATCHING_CONFIG["enable_adaptive_mode"]:
        score = _calculate_adaptive_match_confidence(rec1, rec2)
    else:
        score = _calculate_match_confidence_optimized(rec1, rec2)
    
    # Use baseline threshold by default
    threshold = config.ADAPTIVE_MATCHING_CONFIG["baseline_thresholds"]["strict_threshold"]  # Original 198
    
    # Apply adjustment only when adaptive mode is enabled AND no discriminating fields
    if config.ADAPTIVE_MATCHING_CONFIG["enable_adaptive_mode"]:
        has_birthdate_raw = bool(str(rec1.get("Birthdate", "")).strip() and str(rec2.get("Birthdate", "")).strip())
        has_birthdate_processed = bool(rec1['_opt_bdate_std'] and rec2['_opt_bdate_std'])
        has_sex = bool(str(rec1.get("Sex", "")).strip() and str(rec2.get("Sex", "")).strip())
        
        # Check if we have city data as additional discriminating field
        has_city = bool(str(rec1.get("City", "")).strip() and str(rec2.get("City", "")).strip())
        
        # Only adjust threshold for name-only matching (no birthdate, sex, or city)
        if not (has_birthdate_raw or has_birthdate_processed or has_sex or has_city):
            adjustment = config.ADAPTIVE_MATCHING_CONFIG["threshold_adjustments"]["strict_adjustment"]
            threshold = threshold + adjustment  # 198 + adjustment
    
    return "Exact Match" if score > threshold else "No Match"

def compare_records_standard_configurable(rec1: Dict[str, Any], rec2: Dict[str, Any]) -> str:
    """Configurable comparison: uses enhanced adaptive algorithm when enabled.
    
    Args:
        rec1: First record to compare as a dictionary with string keys
        rec2: Second record to compare as a dictionary with string keys
        
    Returns:
        str: "Fuzzy Match" if score > threshold, otherwise "No Match"
    """
    # Use enhanced adaptive algorithm when adaptive mode is enabled
    if config.ADAPTIVE_MATCHING_CONFIG["enable_adaptive_mode"]:
        score = _calculate_adaptive_match_confidence(rec1, rec2)
    else:
        score = _calculate_match_confidence_optimized(rec1, rec2)
    
    # Use baseline threshold by default
    threshold = config.ADAPTIVE_MATCHING_CONFIG["baseline_thresholds"]["standard_threshold"]  # Your tuned 110
    
    # Apply adjustment only when adaptive mode is enabled AND no discriminating fields
    if config.ADAPTIVE_MATCHING_CONFIG["enable_adaptive_mode"]:
        has_birthdate_raw = bool(str(rec1.get("Birthdate", "")).strip() and str(rec2.get("Birthdate", "")).strip())
        has_birthdate_processed = bool(rec1['_opt_bdate_std'] and rec2['_opt_bdate_std'])
        has_sex = bool(str(rec1.get("Sex", "")).strip() and str(rec2.get("Sex", "")).strip())
        
        # Check if we have city data as additional discriminating field
        has_city = bool(str(rec1.get("City", "")).strip() and str(rec2.get("City", "")).strip())
        
        # Only adjust threshold for name-only matching (no birthdate, sex, or city)
        if not (has_birthdate_raw or has_birthdate_processed or has_sex or has_city):
            adjustment = config.ADAPTIVE_MATCHING_CONFIG["threshold_adjustments"]["standard_adjustment"]
            threshold = threshold + adjustment  # 110 + adjustment
    
    return "Fuzzy Match" if score > threshold else "No Match"

def compare_records_lenient_configurable(rec1: Dict[str, Any], rec2: Dict[str, Any]) -> str:
    """Configurable comparison: uses enhanced adaptive algorithm when enabled.
    
    Args:
        rec1: First record to compare as a dictionary with string keys
        rec2: Second record to compare as a dictionary with string keys
        
    Returns:
        str: "Fuzzy Match" if score > threshold, otherwise "No Match"
    """
    # Use enhanced adaptive algorithm when adaptive mode is enabled
    if config.ADAPTIVE_MATCHING_CONFIG["enable_adaptive_mode"]:
        score = _calculate_adaptive_match_confidence(rec1, rec2)
    else:
        score = _calculate_match_confidence_optimized(rec1, rec2)
    
    # Use baseline threshold by default
    threshold = config.ADAPTIVE_MATCHING_CONFIG["baseline_thresholds"]["lenient_threshold"]  # Original 95
    
    # Apply adjustment only when adaptive mode is enabled AND no discriminating fields
    if config.ADAPTIVE_MATCHING_CONFIG["enable_adaptive_mode"]:
        has_birthdate_raw = bool(str(rec1.get("Birthdate", "")).strip() and str(rec2.get("Birthdate", "")).strip())
        has_birthdate_processed = bool(rec1['_opt_bdate_std'] and rec2['_opt_bdate_std'])
        has_sex = bool(str(rec1.get("Sex", "")).strip() and str(rec2.get("Sex", "")).strip())
        
        # Check if we have city data as additional discriminating field
        has_city = bool(str(rec1.get("City", "")).strip() and str(rec2.get("City", "")).strip())
        
        # Only adjust threshold for name-only matching (no birthdate, sex, or city)
        if not (has_birthdate_raw or has_birthdate_processed or has_sex or has_city):
            adjustment = config.ADAPTIVE_MATCHING_CONFIG["threshold_adjustments"]["lenient_adjustment"]
            threshold = threshold + adjustment  # 95 + adjustment
    
    return "Fuzzy Match" if score > threshold else "No Match"

def _get_blocking_keys_optimized(rec):
    keys = set()
    fname_exp, lname, bdate = rec['_opt_fname_exp'], rec['_opt_lname_raw'], rec['_opt_bdate_std']
    if fname_exp and lname:
        keys.add(f"FL_{fname_exp.upper().replace(' ', '')}_{lname.upper().replace(' ', '')}")
        keys.add(f"SOUNDEX_{jellyfish.soundex(fname_exp)}_{rec['_opt_soundex_lname']}")
    if lname and bdate: keys.add(f"LN_BDATE_{lname.upper()}_{bdate}")
    name_parts = [p for p in [fname_exp, rec['_opt_mname_raw'], lname] if p]
    if len(name_parts) > 1: keys.add(f"SORTED_SOUNDEX_{'_'.join(sorted([jellyfish.soundex(p) for p in name_parts]))}")
    for formal_name in rec['_opt_nickname_set']: keys.add(f"FL_{formal_name.upper().replace(' ', '')}_{lname.upper().replace(' ', '')}")
    return keys

def _generate_pairs_from_blocks(df):
    candidate_pairs, inverted_index = set(), defaultdict(list)
    for i, rec_dict in df.iterrows():
        for key in _get_blocking_keys_optimized(rec_dict): inverted_index[key].append(i)
    for indices in inverted_index.values():
        if len(indices) > 1:
            for pair in itertools.combinations(sorted(indices), 2): candidate_pairs.add(pair)
    return list(candidate_pairs)

def _generate_pairs_from_blocks_2_files(df1, df2):
    candidate_pairs, inverted_index1, inverted_index2 = set(), defaultdict(list), defaultdict(list)
    for i, rec_dict in df1.iterrows():
        for key in _get_blocking_keys_optimized(rec_dict): inverted_index1[key].append(i)
    for i, rec_dict in df2.iterrows():
        for key in _get_blocking_keys_optimized(rec_dict): inverted_index2[key].append(i)
    common_keys = set(inverted_index1.keys()) & set(inverted_index2.keys())
    for key in common_keys:
        for i in inverted_index1[key]:
            for j in inverted_index2[key]: candidate_pairs.add((i, j))
    return list(candidate_pairs)

def process_chunk(chunk, df1_dicts, df2_dicts, comparison_func):
    results = []
    for idx1, idx2 in chunk:
        rec1, rec2 = df1_dicts[idx1], df2_dicts[idx2]
        status = comparison_func(rec1, rec2)
        if status != "No Match": results.append(((idx1, idx2), status))
    return results

def _run_parallel_comparison(df1, df2, comparison_func, candidate_pairs):
    if not candidate_pairs: return []
    df1_dicts = df1.to_dict('index')
    df2_dicts = df2.to_dict('index') if df2 is not None else df1.to_dict('index')
    num_processes = max(1, cpu_count() - 1)
    chunk_size = max(1, len(candidate_pairs) // (num_processes * 4))
    chunks = [candidate_pairs[i:i + chunk_size] for i in range(0, len(candidate_pairs), chunk_size)]
    worker_func = partial(process_chunk, df1_dicts=df1_dicts, df2_dicts=df2_dicts, comparison_func=comparison_func)
    with Pool(processes=num_processes) as pool:
        results = pool.map(worker_func, chunks)
    return [item for sublist in results for item in sublist]

class AnalysisEngine:
    def __init__(self, user_df, master_df, officials_df, nickname_map, user_filepath, province_name, log_callback, status_callback, start_time, final_report_callback, progress_queue):
        self.user_df = user_df
        self.master_df = master_df
        self.officials_df = officials_df
        self.nickname_map = nickname_map
        self.user_filepath = user_filepath
        self.province_name = province_name
        self.log_callback = log_callback
        self.status_callback = status_callback
        self.start_time = start_time
        self.final_report_callback = final_report_callback
        self.progress_queue = progress_queue
        self.symmetrical_map = defaultdict(set)
        self.reports = {}
        self.summary_stats = {}
        self.official_user_indices = set()
        self.linked_user_indices = set()
        self.duplicate_user_indices = set()

    def run_analysis(self):
        if self.user_df is None or self.user_df.empty:
            return

        self._preprocess_data()
        self._perform_matching()
        self._generate_reports()
        self._save_results()

    def _preprocess_data(self):
        self.progress_queue.put(("indeterminate", "Step 1: Preparing and cleaning data..."))
        for nick, formal_list in self.nickname_map.items():
            std_nick, std_formals = nick.lower().replace('.', '').replace(' ', ''), {f.lower().replace('.', '').replace(' ', '') for f in formal_list}
            all_names = {std_nick} | std_formals
            for name in all_names:
                self.symmetrical_map[name].update(all_names)

        self.user_df = _precompute_dataframe(self.user_df, self.symmetrical_map)
        if self.master_df is not None:
            self.master_df = _precompute_dataframe(self.master_df, self.symmetrical_map)
        if self.officials_df is not None:
            self.officials_df = _precompute_dataframe(self.officials_df, self.symmetrical_map)

    def _perform_matching(self):
        self.progress_queue.put(("determinate", 0.1, "Step 2: Analyzing for duplicates and official records..."))
        candidate_pools = {}
        if self.officials_df is not None and not self.officials_df.empty:
            candidate_pools['user_official'] = _generate_pairs_from_blocks_2_files(self.user_df, self.officials_df)
        if self.master_df is not None and not self.master_df.empty:
            candidate_pools['user_master'] = _generate_pairs_from_blocks_2_files(self.user_df, self.master_df)
        candidate_pools['user_user'] = _generate_pairs_from_blocks(self.user_df)

        all_matches, matched_pairs = [], defaultdict(set)
        pass_pipeline = {
            'user_official': [compare_records_strict_configurable, compare_records_standard_configurable, compare_records_lenient_configurable],
            'user_master': [compare_records_strict_configurable, compare_records_standard_configurable],
            'user_user': [compare_records_strict_configurable, compare_records_standard_configurable]
        }

        total_passes = sum(len(funcs) for funcs in pass_pipeline.values())
        current_pass = 0

        for pair_type, funcs in pass_pipeline.items():
            df1_prefix, df2_prefix = pair_type.split('_')
            df1 = self.user_df if df1_prefix == "user" else self.officials_df
            df2 = {"official": self.officials_df, "master": self.master_df, "user": None}[df2_prefix]
            if df1 is None or (df2_prefix != "user" and (df2 is None or df2.empty)):
                current_pass += len(funcs)
                continue
            for func in funcs:
                current_pass += 1
                progress = 0.1 + (0.6 * (current_pass / total_passes))
                self.progress_queue.put(("determinate", progress, f"Step 2: Comparing records ({pair_type})..."))

                pairs_to_check = [p for p in candidate_pools.get(pair_type, []) if tuple(sorted(p)) not in matched_pairs[pair_type]]
                if not pairs_to_check:
                    continue
                pass_results = _run_parallel_comparison(df1, df2, func, pairs_to_check)
                for (i, j), status in pass_results:
                    all_matches.append((f"{df1_prefix}_{i}", f"{df2_prefix if df2 is not None else df1_prefix}_{j}", status))
                    matched_pairs[pair_type].add(tuple(sorted((i, j))))
        self.all_matches = all_matches

    def _generate_reports(self):
        all_nodes = [f"user_{i}" for i in self.user_df.index]
        if self.master_df is not None and not self.master_df.empty:
            all_nodes.extend([f"master_{i}" for i in self.master_df.index])
        if self.officials_df is not None and not self.officials_df.empty:
            all_nodes.extend([f"official_{i}" for i in self.officials_df.index])

        parent = {node: node for node in all_nodes}
        def find(i):
            if parent.get(i, i) == i:
                return i
            parent[i] = find(parent[i])
            return parent[i]

        def union(i, j):
            root_i, root_j = find(i), find(j)
            if root_i != root_j:
                parent[root_j] = root_i

        for node1, node2, _ in self.all_matches:
            union(node1, node2)

        final_groups_dict = defaultdict(list)
        for node in all_nodes:
            final_groups_dict[find(node)].append(node)

        user_involved_groups = {k: v for k, v in final_groups_dict.items() if any(n.startswith("user_") for n in v) and len(v) > 1}

        official_rows, linking_rows, dedupe_rows = [], [], []
        group_id_counters = {"official": 1, "linking": 1, "dedupe": 1}
        group_remarks = {k: "Fuzzy Match" if any(s == "Fuzzy Match" for n1, n2, s in self.all_matches if find(n1) == k and find(n2) == k) else "Exact Match" for k in user_involved_groups}
        sorted_groups = sorted(user_involved_groups.items(), key=lambda item: (0 if any(n.startswith("official_") for n in item[1]) else (1 if any(n.startswith("master_") for n in item[1]) else 2), 0 if group_remarks[item[0]] == "Exact Match" else 1, sorted([int(n.split('_')[1]) for n in item[1] if n.startswith("user_")])[0]))
        
        processed_user_indices = set()
        for group_key, group_members in sorted_groups:
            user_nodes = {int(n.split('_')[1]) for n in group_members if n.startswith("user_")}
            if not user_nodes.isdisjoint(processed_user_indices):
                continue
            
            has_official = any(n.startswith("official_") for n in group_members)
            has_master = any(n.startswith("master_") for n in group_members)

            if has_official:
                for node in sorted(group_members):
                    source, idx_str = node.split('_'); idx = int(idx_str)
                    df = self.user_df if source == 'user' else (self.officials_df if source == 'official' else None)
                    if df is not None:
                        official_rows.append({**df.loc[idx].to_dict(), "group_id": group_id_counters["official"], "Row": f"{'userfile' if source == 'user' else 'official'} {idx + 2}", "Remarks": "Official"})
                self.official_user_indices.update(user_nodes)
                group_id_counters["official"] += 1
            elif has_master:
                for node in sorted(group_members):
                    source, idx_str = node.split('_'); idx = int(idx_str)
                    df = self.user_df if source == 'user' else (self.master_df if source == 'master' else None)
                    if df is not None:
                        linking_rows.append({**df.loc[idx].to_dict(), "group_id": group_id_counters["linking"], "Row": f"{'userfile' if source == 'user' else 'masterdb'} {idx + 2}", "Remarks": group_remarks[group_key]})
                self.linked_user_indices.update(user_nodes)
                group_id_counters["linking"] += 1
            elif len(user_nodes) > 1:
                for node in sorted(group_members):
                    source, idx_str = node.split('_'); idx = int(idx_str)
                    if source == 'user':
                        dedupe_rows.append({**self.user_df.loc[idx].to_dict(), "group_id": group_id_counters["dedupe"], "Row": f"userfile {idx + 2}", "Remarks": group_remarks[group_key]})
                self.duplicate_user_indices.update(user_nodes)
                group_id_counters["dedupe"] += 1
            
            processed_user_indices.update(user_nodes)

        self.reports["officials"], self.reports["linking"], self.reports["dedupe"] = pd.DataFrame(official_rows), pd.DataFrame(linking_rows), pd.DataFrame(dedupe_rows)

    def _save_results(self):
        self.progress_queue.put(("determinate", 0.8, "Step 3: Compiling findings into report file..."))
        
        opt_cols_to_drop = [c for c in self.user_df.columns if c.startswith('_opt_')]
        self.user_df.drop(columns=opt_cols_to_drop, inplace=True, errors='ignore')
        
        for key, df in self.reports.items():
            if not df.empty:
                opt_cols = [c for c in df.columns if c.startswith('_opt_')]
                df.drop(columns=opt_cols, inplace=True, errors='ignore')
                df.sort_values(by=['group_id', 'Row'], inplace=True)
                df['group_id'] = (df['group_id'] != df['group_id'].shift()).cumsum()
                # Base ordering starts with group and row
                base_cols = ["group_id", "Row"]
                # For linking section, insert optional informational 'Batch Name' immediately after 'Row' if present
                if key == "linking" and "Batch Name" in df.columns:
                    base_cols.append("Batch Name")
                # Then add the core intended columns (names, birthdate, etc.)
                base_cols.extend([c for c in INTENDED_COLS if c in df.columns])
                # Officials have extra location fields before remarks
                order_list = base_cols + ["Position", "Barangay", "Remarks"] if key == "officials" else base_cols + ["Remarks"]
                existing_cols = [col for col in order_list if col in df.columns]
                other_cols = sorted([col for col in df.columns if col not in existing_cols])
                self.reports[key] = df[existing_cols + other_cols]

        total_user_records = len(self.user_df)
        official_count = len(self.official_user_indices)
        linking_only_indices = self.linked_user_indices - self.official_user_indices
        linking_count = len(linking_only_indices)
        duplicate_only_indices = self.duplicate_user_indices - self.official_user_indices - linking_only_indices
        duplicate_count = len(duplicate_only_indices)
        unique_count = max(0, total_user_records - (official_count + linking_count + duplicate_count))

        self.summary_stats = {
            "officials": official_count,
            "linking": linking_count,
            "duplicates": duplicate_count,
            "unique": unique_count,
            "total": total_user_records
        }
        
        end_time = datetime.now()

        if config.REPORT_FORMAT == 'PDF':
            output_dir = os.path.dirname(self.user_filepath)
            base_name = os.path.splitext(os.path.basename(self.user_filepath))[0]
            province_tag = self.province_name.replace(" ", "_")
            output_base = os.path.join(output_dir, f"{base_name}_{province_tag}_report")
            counter = 1
            while os.path.exists(f"{output_base}_{counter}.pdf") or os.path.exists(f"{output_base}_{counter}.xlsx"):
                counter += 1
            pdf_output_path = f"{output_base}_{counter}.pdf"
            excel_output_path = f"{output_base}_{counter}.xlsx"
        else:
            excel_output_path = get_automatic_output_path(self.user_filepath, self.province_name, is_pdf=False)
            pdf_output_path = None

        final_output_path = excel_output_path
        
        excel_generation_success = generate_excel_report(self.reports, self.user_df, self.user_filepath, excel_output_path, self.start_time, end_time, self.summary_stats, self.master_df, self.officials_df, self.official_user_indices, (self.linked_user_indices | self.duplicate_user_indices))
        
        if not excel_generation_success:
            self.log_callback(f"❌ FILE LOCKED: Could not save report to {excel_output_path}.")
            self.status_callback("main", "Error: Report file is locked.", "error")
            return

        if config.REPORT_FORMAT == 'PDF':
            conversion_success = excel_converter.convert_to_pdf(excel_output_path, pdf_output_path)
            if conversion_success:
                self.log_callback("✅ PDF conversion successful.")
                final_output_path = pdf_output_path
                try:
                    os.remove(excel_output_path)
                except OSError:
                    self.log_callback(f"⚠️ Could not remove temporary Excel file at: {excel_output_path}")
            else:
                self.log_callback("⚠️ PDF conversion failed. Please ensure Microsoft Excel is installed and licensed.")
                self.log_callback(f"✅ Saving as Excel report instead at: {excel_output_path}")

        self.progress_queue.put(("determinate", 1.0, "Step 4: Finalizing analysis..."))
        summary_table_data = create_summary_section(self.summary_stats, self.user_df, self.master_df, self.officials_df)
        summary_lines = [f"--- Summary Report ---\nScan Started:   {self.start_time:%Y-%m-%d %H:%M:%S}\nScan Finished:  {end_time:%Y-%m-%d %H:%M:%S}\nTotal Duration: {format_duration((end_time - self.start_time).total_seconds())}\n"]
        if summary_table_data:
            headers = ["Analysis Section", "Total Scanned", "Names Found"]
            col_widths = [max(len(str(h)), max((len(str(row[i])) for row in summary_table_data), default=0)) for i, h in enumerate(headers)]
            header_line = " | ".join([str(h).ljust(w) for h, w in zip(headers, col_widths)])
            separator_line = '-' * len(header_line)
            summary_lines.extend([separator_line, header_line, separator_line])
            for row in summary_table_data:
                if row[0] == "Total":
                    summary_lines.extend([separator_line, f"{str(row[0]).ljust(col_widths[0])} | {' '.rjust(col_widths[1])} | {str(row[2]).rjust(col_widths[2])}", separator_line])
                else:
                    summary_lines.append(" | ".join([str(cell).ljust(w) if j==0 else str(cell).rjust(w) for j, (cell, w) in enumerate(zip(row, col_widths))]))
        
        self.log_callback("\n" + "\n".join(summary_lines))
        self.final_report_callback(final_output_path)
        self.status_callback("main", "Analysis complete. Report saved.", "success")

def run_analysis(user_df, master_df, officials_df, nickname_map, user_filepath, province_name, log_callback, status_callback, start_time, final_report_callback, progress_queue):
    engine = AnalysisEngine(user_df, master_df, officials_df, nickname_map, user_filepath, province_name, log_callback, status_callback, start_time, final_report_callback, progress_queue)
    engine.run_analysis()

def generate_excel_report(reports, user_df, user_filepath, output_filename, start_time, end_time, summary_stats, master_df, officials_df, official_user_indices, chart_duplicate_indices):
    if user_df is not None and not user_df.empty: user_df.fillna('', inplace=True)
    for key in reports:
        if reports[key] is not None and not reports[key].empty: reports[key] = reports[key].fillna('')
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    
    summary_table_data = create_summary_section(summary_stats, user_df, master_df, officials_df)
    create_dashboard_sheet(wb.create_sheet("Dashboard", 0), reports, user_df, user_filepath, start_time, end_time, summary_stats, official_user_indices, chart_duplicate_indices)
    create_user_data_sheet(wb.create_sheet("User File Data", 1), user_df, reports, official_user_indices)
    create_analysis_report_sheet(wb.create_sheet("Analysis Report", 2), reports, user_filepath, start_time, end_time, summary_stats, user_df, master_df, officials_df, summary_table_data)
    try:
        wb.save(output_filename)
        return True
    except PermissionError:
        return False

def create_dashboard_sheet(ws, reports, user_df, user_filepath, start_time, end_time, summary_stats, official_user_indices, chart_duplicate_indices):
    total_records, officials_kpi_count, unique_records_count = summary_stats.get("total", 0), summary_stats.get("officials", 0), summary_stats.get("unique", 0)
    duplicates_kpi_count = summary_stats.get("linking", 0) + summary_stats.get("duplicates", 0)
    exact_matches_count, fuzzy_matches_count = 0, 0
    all_duplicates_df = pd.concat([reports.get("linking", pd.DataFrame()), reports.get("dedupe", pd.DataFrame())], ignore_index=True)
    if not all_duplicates_df.empty and chart_duplicate_indices:
        user_rows_in_dupes = all_duplicates_df[all_duplicates_df['Row'].str.startswith('userfile', na=False)].copy()
        if not user_rows_in_dupes.empty:
            user_rows_in_dupes['original_index'] = user_rows_in_dupes['Row'].str.extract(r'(\d+)$', expand=False).astype(int) - 2
            chart_records = user_rows_in_dupes[user_rows_in_dupes['original_index'].isin(chart_duplicate_indices)]
            if not chart_records.empty:
                group_remark_map = all_duplicates_df.drop_duplicates(subset=['group_id']).set_index('group_id')['Remarks']
                unique_chart_records = chart_records.drop_duplicates(subset=['original_index'])
                unique_chart_records['remark_type'] = unique_chart_records['group_id'].map(group_remark_map)
                remark_counts = unique_chart_records['remark_type'].value_counts()
                exact_matches_count, fuzzy_matches_count = remark_counts.get("Exact Match", 0), remark_counts.get("Fuzzy Match", 0)
    top_cities = pd.Series(dtype=object)
    all_findings_df = pd.concat([reports.get("officials"), reports.get("linking"), reports.get("dedupe")], ignore_index=True)
    if not all_findings_df.empty:
        city_col = 'City' if 'City' in all_findings_df.columns else None
        if city_col:
            user_records_in_findings = all_findings_df[all_findings_df['Row'].str.startswith('userfile', na=False)]
            if not user_records_in_findings.empty:
                top_cities = user_records_in_findings[city_col].fillna('Unknown').value_counts().nlargest(5)
    kpi_value_font, kpi_title_font, chart_header_font = Font(name='Calibri', size=26, bold=True), Font(name='Calibri', size=12, bold=True), Font(name='Calibri', size=12, bold=True)
    header_fill, chart_card_fill, background_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"), PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thick_side = Side(style='thick', color="000000")
    ws.sheet_view.showGridLines = False
    for col in ws.columns:
        for cell in col: cell.fill = background_fill
    ws.column_dimensions['A'].width, ws.column_dimensions['G'].width, ws.column_dimensions['M'].width, ws.column_dimensions['S'].width = 2, 2, 2, 2
    for col in ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J', 'K', 'L', 'N', 'O', 'P', 'Q', 'R']: ws.column_dimensions[col].width = 10
    ws.merge_cells("B1:R1")
    title_cell = ws['B1']; title_cell.value, title_cell.font, title_cell.alignment = f"Analysis Dashboard for: {os.path.basename(user_filepath)}", Font(name='Calibri', size=16, bold=True), Alignment(horizontal='left', vertical='center')
    kpi_data = [("B3:F5", "Total Records", total_records), ("H3:L5", "Officials Found", officials_kpi_count), ("N3:R5", "Duplicates Found", duplicates_kpi_count)]
    for area, title, value in kpi_data:
        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(area)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c); cell.fill = header_fill
                cell.border = Border(left=thick_side if c == min_col else None, right=thick_side if c == max_col else None, top=thick_side if r == min_row else None, bottom=thick_side if r == max_row else None)
        ws.merge_cells(f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{min_row}")
        ws.merge_cells(f"{get_column_letter(min_col)}{min_row + 1}:{get_column_letter(max_col)}{max_row}")
        title_cell, value_cell = ws.cell(row=min_row, column=min_col), ws.cell(row=min_row + 1, column=min_col)
        title_cell.value, title_cell.font, title_cell.alignment = title, kpi_title_font, Alignment(horizontal='center', vertical='center')
        value_cell.value, value_cell.font, value_cell.alignment = value, kpi_value_font, Alignment(horizontal='center', vertical='center')
    chart_card_data = [("B7:F7", "B8:F19", "Top Cities with Findings"), ("H7:L7", "H8:L19", "Duplicate Match Quality"), ("N7:R7", "N8:R19", "File Composition")]
    for header_area, body_area, title in chart_card_data:
        ws.merge_cells(header_area)
        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(header_area)
        title_cell = ws.cell(row=min_row, column=min_col); title_cell.value, title_cell.font, title_cell.alignment, title_cell.fill = title, chart_header_font, Alignment(horizontal='center', vertical='center'), header_fill
        for c in range(min_col, max_col + 1): ws.cell(row=min_row, column=c).border = Border(left=thick_side if c == min_col else None, right=thick_side if c == max_col else None, top=thick_side, bottom=thick_side)
        min_col_b, min_row_b, max_col_b, max_row_b = openpyxl.utils.cell.range_boundaries(body_area)
        for r in range(min_row_b, max_row_b + 1):
            for c in range(min_col_b, max_col_b + 1): ws.cell(row=r, column=c).fill = chart_card_fill
    chart_data_ws = ws.parent.create_sheet("ChartData")
    chart_data_ws['A1'], chart_data_ws['B1'] = "Category", ""  # Remove 'Count' label
    # Prepare the actual data values
    actual_values = {
        "Unique Records": unique_records_count,
        "Duplicate Records": duplicates_kpi_count,
        "Official Records": officials_kpi_count
    }
    
    # Calculate total from actual values
    total = sum(actual_values.values())
    
    # Prepare display data using ACTUAL counts to preserve accurate percentages
    display_data = []
    for label, count in actual_values.items():
        display_count = count
        display_data.append((label, display_count, count))  # Store both display and actual count
    
    # If one category is 100%, only show that one
    if total > 0 and any(count == total for count in actual_values.values()):
        doughnut_data = [(label, display_count) for label, display_count, actual in display_data if actual > 0]
    else:
        doughnut_data = [(label, display_count) for label, display_count, _ in display_data]
    
    # Write the display data to the worksheet
    for i, (label, display_count) in enumerate(doughnut_data, 2):
        chart_data_ws[f'A{i}'], chart_data_ws[f'B{i}'] = label, display_count
    chart_data_ws['D1'], chart_data_ws['E1'] = "Match Type", "Count"
    bar_data = [("Fuzzy", fuzzy_matches_count), ("Exact", exact_matches_count)]; filtered_bar_data = [item for item in bar_data if item[1] > 0]
    if filtered_bar_data:
        for i, (label, value) in enumerate(filtered_bar_data, 2): chart_data_ws[f'D{i}'], chart_data_ws[f'E{i}'] = label, value
    chart_data_ws['G1'], chart_data_ws['H1'] = "City", "Count"
    if not top_cities.empty:
        for i, (city, count) in enumerate(top_cities.items(), 2): chart_data_ws[f'G{i}'], chart_data_ws[f'H{i}'] = city, count
    chart_data_ws.sheet_state = 'hidden'
    def add_empty_chart_message(cell_range, message): ws.merge_cells(cell_range); cell = ws[cell_range.split(':')[0]]; cell.value, cell.font, cell.alignment = message, Font(name='Calibri', size=14, bold=True, color="808080"), Alignment(horizontal='center', vertical='center')
    if not top_cities.empty:
        chart1 = BarChart()
        chart1.type, chart1.title, chart1.legend, chart1.dataLabels, chart1.varyColors, chart1.height, chart1.width, chart1.y_axis.majorGridlines = "col", None, None, DataLabelList(showSerName=False, showCatName=False, showVal=True), True, 6.15, 9.7, None
        chart1.dataLabels.showLegendKey = False
        chart1.x_axis.title, chart1.y_axis.title = None, None
        data, cats = Reference(chart_data_ws, min_col=8, min_row=1, max_row=1 + len(top_cities)), Reference(chart_data_ws, min_col=7, min_row=2, max_row=1 + len(top_cities))
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        for s in chart1.series: s.title = None
        ws.add_chart(chart1, "B8")
    else: add_empty_chart_message("B8:F19", "No Findings in Cities")
    if (exact_matches_count + fuzzy_matches_count) > 0 and filtered_bar_data:
        chart2 = BarChart()
        chart2.type, chart2.title, chart2.legend, chart2.dataLabels, chart2.varyColors, chart2.height, chart2.width = "bar", None, None, DataLabelList(showSerName=False, showCatName=True, showVal=True), True, 6.15, 9.7
        chart2.x_axis.graphicalProperties, chart2.dataLabels.showLegendKey, chart2.y_axis.majorGridlines = GraphicalProperties(ln=LineProperties(noFill=True)), False, None
        chart2.x_axis.title, chart2.y_axis.title = None, None
        data, cats = Reference(chart_data_ws, min_col=5, min_row=1, max_row=1 + len(filtered_bar_data)), Reference(chart_data_ws, min_col=4, min_row=2, max_row=1 + len(filtered_bar_data))
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        for s in chart2.series: s.title = None
        ws.add_chart(chart2, "H8")
    else: add_empty_chart_message("H8:L19", "No Duplicates Found")
    if total_records > 0:
        # Calculate the total and original percentages for each category (for explosion logic only)
        total = sum(count for _, _, count in display_data)
        percentages = [((count / total) * 100 if total > 0 else 0) for _, _, count in display_data]

        # Always write ORIGINAL counts to preserve accurate percentages on labels
        for i, (_, display_count, _) in enumerate(display_data):
            chart_data_ws[f'B{i+2}'] = display_count

        # Create PIE chart or DOUGHNUT chart
        chart = PieChart()
        chart.title = None
        chart.legend.position = 'b'
        data = Reference(chart_data_ws, min_col=2, min_row=2, max_row=1 + len(doughnut_data))
        cats = Reference(chart_data_ws, min_col=1, min_row=2, max_row=1 + len(doughnut_data))
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.varyColors = True
        chart.height = 6.15
        chart.width = 9.7
        
        # --- FINAL LOGIC FOR SLICE EXPLOSION & LABELS ---
        
        # 1. Access the chart's series. Pie charts have only one.
        pie_series = chart.series[0]

        # 2. Create a list of DataPoint objects. We will customize only the ones that need exploding.
        # The 'explosion' value is a percentage (e.g., 25 means pushed out by 25%).
        points = []
        EXPLOSION_THRESHOLD_PERCENT = 10.0 # Slices smaller than this get detached
        for i, p in enumerate(percentages):
            if 0 < p < EXPLOSION_THRESHOLD_PERCENT:
                # This slice is small, so create a DataPoint to explode it.
                pt = DataPoint(idx=i, explosion=25) 
                points.append(pt)
        
        # 3. Assign the list of custom DataPoints to the series.
        pie_series.dPt = points

        # 4. Configure the data labels. We don't force a position, allowing
        # Excel to intelligently place them inside or outside based on the space
        # created by the explosion.
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showLeaderLines = True # Ensure lines are drawn for outside labels
        # NO dLblPos setting - let Excel decide based on available space.
        chart.dataLabels.showCategoryName = False
        chart.dataLabels.showLegendKey = False
        # Show decimals for small percentages (e.g., 0.05%)
        try:
            chart.dataLabels.numFmt = "0.00%"
        except Exception:
            # If the runtime openpyxl version doesn't support numFmt here, fail silently
            pass
        
        # Add the chart to the worksheet
        ws.add_chart(chart, "N8")
    elif total_records == 0: 
        add_empty_chart_message("N8:R19", "No Records to Analyze")

def create_user_data_sheet(ws, user_df, reports, official_user_indices):
    official_fill, exact_fill, fuzzy_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"), PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"); bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Track the 'Sex' column letter so we can force width at the end
    sex_col_letter_user_sheet = None
    exact_dupe_indices, fuzzy_dupe_indices = set(), set()
    all_duplicates = pd.concat([reports.get("dedupe", pd.DataFrame()), reports.get("linking", pd.DataFrame())], ignore_index=True)
    if not all_duplicates.empty:
        for _, row in all_duplicates.iterrows():
            if 'userfile' not in str(row.get('Row', '')): continue
            match = re.search(r'(\d+)$', str(row.get('Row', '')))
            if match:
                original_index = int(match.group(1)) - 2
                if original_index >= 0:
                    if "Exact Match" in str(row.get('Remarks', '')): exact_dupe_indices.add(original_index)
                    elif "Fuzzy Match" in str(row.get('Remarks', '')): fuzzy_dupe_indices.add(original_index)
    if user_df is not None:
        df_to_write = user_df.copy()
        df_to_write['Legend'] = ''
        if official_user_indices: df_to_write.loc[list(official_user_indices), 'Legend'] = 'Official'
        if exact_dupe_indices: df_to_write.loc[list(exact_dupe_indices), 'Legend'] = 'Exact Match'
        if fuzzy_dupe_indices: df_to_write.loc[list(fuzzy_dupe_indices), 'Legend'] = 'Fuzzy Match'
        headers = list(df_to_write.columns)
        for c_idx, value in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=value)
            cell.font = bold_font
            cell.border = thin_border
            # Remember the column letter for the 'Sex' header
            if str(value).strip().lower() == "sex":
                sex_col_letter_user_sheet = get_column_letter(c_idx)
        if not df_to_write.empty:
            for r_idx, row_data in df_to_write.iterrows():
                fill = None
                if r_idx in official_user_indices: fill = official_fill
                elif r_idx in exact_dupe_indices: fill = exact_fill
                elif r_idx in fuzzy_dupe_indices: fill = fuzzy_fill
                for c_idx, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=r_idx + 2, column=c_idx, value=cell_value)
                    if fill: cell.fill = fill
                    cell.border = thin_border
            contact_col_name = next((c for c in ["Contact Number", "ContactNumber"] if c in headers), None)
            if contact_col_name:
                contact_col_idx = headers.index(contact_col_name) + 1
                for row_num in range(2, ws.max_row + 1): ws.cell(row=row_num, column=contact_col_idx).number_format = '@'
    if official_user_indices or exact_dupe_indices or fuzzy_dupe_indices:
        legend_start_row = ws.max_row + 2
        ws.cell(row=legend_start_row, column=1, value="Legend").font = bold_font
        legend_data, thin_border = [("Government Official", official_fill), ("Duplicate (Exact Match)", exact_fill), ("Duplicate (Fuzzy Match)", fuzzy_fill)], Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for i, (text, fill) in enumerate(legend_data):
            current_row = legend_start_row + 1 + i
            color_cell = ws.cell(row=current_row, column=1); color_cell.fill, color_cell.border = fill, thin_border
            text_cell = ws.cell(row=current_row, column=2, value=text); text_cell.border = thin_border
        # Add the uncolored Unique Names entry last, as requested
        unique_row = legend_start_row + 1 + len(legend_data)
        color_cell = ws.cell(row=unique_row, column=1)
        color_cell.border = thin_border
        ws.cell(row=unique_row, column=2, value="Unique Names").border = thin_border
    else:
        # No colored categories exist. Still render the legend with only the uncolored Unique Names row.
        legend_start_row = ws.max_row + 2
        ws.cell(row=legend_start_row, column=1, value="Legend").font = bold_font
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        unique_row = legend_start_row + 1
        color_cell = ws.cell(row=unique_row, column=1)
        color_cell.border = thin_border
        ws.cell(row=unique_row, column=2, value="Unique Names").border = thin_border
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 0.5, 50)
    # Enforce 'Sex' column width to 7 after auto-fit
    if sex_col_letter_user_sheet:
        ws.column_dimensions[sex_col_letter_user_sheet].width = 7

def create_analysis_report_sheet(ws, reports, user_filepath, start_time, end_time, summary_stats, user_df, master_df, officials_df, summary_table_data):
    current_row, user_filename, center_align = 1, os.path.basename(user_filepath), Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Track section ranges to apply custom column width rules later
    section_ranges = {
        "officials": None,
        "linking": None,
        "dedupe": None,
    }
    
    def write_df_to_sheet(df, title):
        nonlocal current_row
        if df is None or df.empty: ws.cell(row=current_row, column=1, value=f"--- {title} ---").font = Font(bold=True); current_row += 1; ws.cell(row=current_row, column=1, value="[No Records Found]"); current_row += 2; return
        ws.cell(row=current_row, column=1, value=f"--- {title} ---").font = Font(bold=True); current_row += 1
        headers = list(df.columns)
        header_row = current_row
        start_row_for_df = current_row + 1
        
        # Write headers with borders
        for c_idx, col_name in enumerate(headers, 1): 
            cell = ws.cell(row=current_row, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = center_align if "group_id" in col_name.lower() else None
        
        # Write data with borders
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row_for_df):
            for c_idx, value in enumerate(row, 1): 
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align if c_idx <= len(headers) and "group_id" in headers[c_idx-1].lower() else None
        
        current_row += len(df)
        
        # Auto-fit columns B to M for Officials Found section (even if no records)
        if "Officials Found" in title:
            for col_idx in range(2, 14):  # Columns B to M (2 to 13)
                max_length = 0
                # Consider any header we may have written for this section
                header_cell = ws.cell(row=current_row - len(df), column=col_idx)
                if header_cell.value:
                    max_length = max(max_length, len(str(header_cell.value)))
                # Consider all values written for this section
                for row_idx in range(start_row_for_df, start_row_for_df + len(df)):
                    data_cell = ws.cell(row=row_idx, column=col_idx)
                    if data_cell.value is not None:
                        max_length = max(max_length, len(str(data_cell.value)))
                # Use tighter heuristic similar to Excel: +0.25 char padding, 50 cap
                adjusted_width = max(ws.column_dimensions[get_column_letter(col_idx)].width or 0, min(max_length + 0.25, 50))
                if adjusted_width == 0: adjusted_width = min(max_length + 0.25, 50)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Format contact number column
        contact_col_name = next((c for c in ["Contact Number", "ContactNumber"] if c in headers), None)
        if contact_col_name:
            contact_col_idx = headers.index(contact_col_name) + 1
            for row_num in range(start_row_for_df, current_row + 1): ws.cell(row=row_num, column=contact_col_idx).number_format = '@'

        # Record section ranges for custom width rules
        section_key = None
        if "Officials Found" in title:
            section_key = "officials"
        elif "Linked Records" in title:
            section_key = "linking"
        elif "Duplicates Found" in title:
            section_key = "dedupe"
        if section_key is not None:
            section_ranges[section_key] = {
                "header_row": header_row,
                "data_start": start_row_for_df,
                "data_end": (start_row_for_df + len(df) - 1)
            }
        current_row += 2
    
    write_df_to_sheet(reports.get("officials"), f"Officials Found in User File ({user_filename} vs OfficialsDB)")
    write_df_to_sheet(reports.get("linking"), f"Linked Records ({user_filename} + MasterDB)")
    write_df_to_sheet(reports.get("dedupe"), f"Duplicates Found in User File ({user_filename})")
    
    ws.cell(row=current_row, column=1, value="--- Summary Report ---").font = Font(bold=True); current_row += 1
    ws.cell(row=current_row, column=1, value="Scan Started:"), ws.cell(row=current_row, column=2, value=start_time.strftime("%Y-%m-%d %H:%M:%S")); current_row += 1
    ws.cell(row=current_row, column=1, value="Scan Finished:"), ws.cell(row=current_row, column=2, value=end_time.strftime("%Y-%m-%d %H:%M:%S")); current_row += 1
    ws.cell(row=current_row, column=1, value="Total Duration:"), ws.cell(row=current_row, column=2, value=format_duration((end_time - start_time).total_seconds())); current_row += 2
    
    if summary_table_data:
        headers = ["Analysis Section", "Total Scanned", "Names Found"]
        start_table_row = current_row
        for c_idx, header_val in enumerate(headers, 1): ws.cell(row=start_table_row, column=c_idx, value=header_val)
        for r_idx, data_row in enumerate(summary_table_data, 1):
            for c_idx, cell_val in enumerate(data_row, 1): ws.cell(row=start_table_row + r_idx, column=c_idx, value=cell_val)
        bold_font, thin_border = Font(bold=True), Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row_idx in range(start_table_row, start_table_row + len(summary_table_data) + 1):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx); cell.border = thin_border
                if row_idx == start_table_row or ws.cell(row=row_idx, column=1).value == "Total": cell.font = bold_font
                cell.alignment = Alignment(horizontal='left') if col_idx == 1 else Alignment(horizontal='right')
        col_widths = [max(len(str(h)), max((len(str(row[i])) for row in summary_table_data), default=0)) for i, h in enumerate(headers)]
        for i, width in enumerate(col_widths): ws.column_dimensions[get_column_letter(i + 1)].width = width + 5

    # Ensure columns B to M are auto-fitted across the sheet (acts like doing this once on Officials)
    for col_idx in range(2, 14):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            cell_val = row[0].value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        if max_len:
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 0.25, 50)

    # Set column A to constant width of 26 (requested)
    ws.column_dimensions['A'].width = 26

    # Apply custom column width rules per your request:
    # - Column I (9): width based on Linkage section content (City)
    # - Column J (10): width based on Officials section content (Contact Number)
    # - Column K (11): width based on Officials section content (Position)
    def _autosize_by_range(col_idx: int, start_row: int, end_row: int, pad: float = 0.25):
        if start_row is None or end_row is None or start_row <= 0 or end_row <= 0:
            return
        max_len = 0
        # Include header row as well
        header_row = start_row - 1
        for r in [header_row] + list(range(start_row, end_row + 1)):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        if max_len:
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + pad, 50)

    if section_ranges["linking"]:
        rng = section_ranges["linking"]
        # Add extra padding for longer city names like 'Mandaluyong' so letters aren't clipped
        _autosize_by_range(9, rng["data_start"], rng["data_end"], pad=2.0)   # Column I
    if section_ranges["officials"]:
        rng = section_ranges["officials"]
        _autosize_by_range(10, rng["data_start"], rng["data_end"])  # Column J
        _autosize_by_range(11, rng["data_start"], rng["data_end"])  # Column K